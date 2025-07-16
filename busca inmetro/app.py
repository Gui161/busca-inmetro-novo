import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl import load_workbook

#Criar planilha
equipamentos = []

planilha = load_workbook(r"busca inmetro\entradas\dados.xlsx")

aba_ativa = planilha["Sheet1"]
for celula in aba_ativa["A"]:
    
    equipamentos.append(celula.value)

print(equipamentos)

# Inicializar o navegador
navegador = webdriver.Chrome()
wait = WebDriverWait(navegador, 10)
navegador.get("http://www.inmetro.gov.br/laboratorios/rbc/consulta.asp")
navegador.maximize_window()

# Criar planilha excel
plan_saida = openpyxl.Workbook()
pagina_resultados = plan_saida["Sheet"]
pagina_resultados.title = "Resultados"
pagina_resultados.append(["Serviço", "Quantidade de equipamentos"])


try:
    for equipamento in equipamentos:
        
        equipamento_formatado = equipamento.replace("/", "-")
        plan_saida.create_sheet(equipamento_formatado)
        pagina_atual = plan_saida[equipamento_formatado]
        pagina_atual.append(["Número de acreditação", "Nome da empresa", "Status", "Estado", "Area", "Número de empresas que realizam a calibração de {}".format(equipamento)])
        contador = 0
        print(f"Pesquisando por: {equipamento}\n")

        # Localizar e preencher o campo de pesquisa
        campo_nome = wait.until(EC.presence_of_element_located((By.NAME, "nom_servico")))
        campo_nome.clear()
        campo_nome.send_keys(equipamento)

        # Clicar no botão de pesquisar
        botao_pesquisar = navegador.find_element(By.NAME, "Submit")
        botao_pesquisar.click()

        pagina = 1
        contador = 0
        while True:
            try:
                # Esperar a tabela de resultados carregar
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, "listagem")))
                empresas = navegador.find_elements(By.CLASS_NAME, "listagem")
                lista_de_empresas = []

                # Coletar dados das empresas
                for i in range(0, len(empresas), 6):
                    if i + 5 < len(empresas):
                        empresa = {
                            "ID": empresas[i].text,
                            "Nome": empresas[i + 1].text,
                            "Status": empresas[i + 2].text,
                            "Estado": empresas[i + 3].text,
                            "Categoria": empresas[i + 4].text,
                        }
                        lista_de_empresas.append(empresa)

                # Exibir resultados da página atual
                if lista_de_empresas:
                    for empresa in lista_de_empresas:
                        
                        pagina_atual.append([empresa["ID"], empresa["Nome"], empresa["Status"], empresa["Estado"], empresa["Categoria"]])

                        contador +=1
                    
                else:
                    print("Nenhuma empresa encontrada para este equipamento.")
                    break

                # Procurar botão de próxima página apenas depois de exibir os resultados
                botoes = navegador.find_elements(By.TAG_NAME, "a")
                proxima_pagina = None

                for botao in botoes:
                    try:
                        if botao.text.isdigit() and int(botao.text) == pagina + 1:
                            proxima_pagina = botao
                            break
                    except ValueError:
                        continue

                if proxima_pagina:
                    print(f"Indo para a página {pagina + 1}\n")
                    proxima_pagina.click()
                    pagina += 1
                    time.sleep(3)
                else:
                    print("Fim dos resultados.\n")
                    break
            except Exception as e:
                print(f"Erro ao processar página {pagina}: {e}")
                break
        pagina_atual.cell(row=2, column=6, value=contador) 
        
        
        pagina_resultados.append([equipamento, contador])
        plan_saida.save("dados_saida.xlsx")
        print("Foram encontradas {} empresas que realizam a calibração de {}".format(contador, equipamento))
        time.sleep(2)       
        botao_voltar = navegador.find_element(By.XPATH, "/html/body/table[2]/tbody/tr[2]/td/a[2]/img").click()
        
        
except Exception as e:
    print(f"Erro geral: {e}")
finally:
    
    
    
    
    time.sleep(5)
    navegador.quit()